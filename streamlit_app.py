import streamlit as st
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objs as go 
import json
import ydata_profiling
from openpyxl import load_workbook
from streamlit_pandas_profiling import st_profile_report
st.set_option('deprecation.showPyplotGlobalUse', False)

st.markdown("<h1 style='text-align: center; color: white;'>Automated EDA Using StreamLit</h1>", unsafe_allow_html=True)
st.markdown("<div id='linkto_top'></div>", unsafe_allow_html=True)

######Extra functions
def datf_inf(df):
    df.columns = df.columns.str.replace(' ', '_')
    buffer = io.StringIO()
    df.info(buf=buffer)
    s = buffer.getvalue()

    df_info = s.split('\n')

    counts = []
    names = []
    nn_count = []
    dtype = []
    for i in range(5, len(df_info) - 3):
        line = df_info[i].split()
        counts.append(line[0])
        names.append(line[1])
        nn_count.append(line[2])
        dtype.append(line[4])

    df_info_dataframe = pd.DataFrame(
        data={'#': counts, 'Column': names, 'Non-Null Count': nn_count, 'Data Type': dtype})
    return df_info_dataframe.drop('#', axis=1)

def sidebar_multiselect_container(massage, arr, key):
    container = st.sidebar.container()
    select_all_button = st.sidebar.checkbox("Select all for " + key + " plots")
    if select_all_button:
        selected_num_cols = container.multiselect(massage, arr, default=list(arr))
    else:
        selected_num_cols = container.multiselect(massage, arr, default=arr[0])

    return selected_num_cols
def datf_nval(df):
    res = pd.DataFrame(df.isnull().sum()).reset_index()
    res['Percentage'] = round(res[0] / df.shape[0] * 100, 2)
    res['Percentage'] = res['Percentage'].astype(str) + '%'
    return res.rename(columns={'index': 'Column', 0: 'Number of null values'})
##############


# Function to load data from different file formats
def load_data(file_path, file_format, sheet_name=None):
    if file_format == 'CSV':
        return pd.read_csv(file_path)
    elif file_format in ['XLS', 'XLSX']:
        if sheet_name is None:
            # Detect sheet names without reading the data
            wb = load_workbook(file_path, read_only=True)
            return wb.sheetnames
        else:
            xls = pd.ExcelFile(file_path)
            return xls.parse(sheet_name)
    elif file_format == 'JSON':
        return pd.read_json(file_path)
    elif file_format == 'TSV':
        return pd.read_csv(file_path, sep='\t')
    else:
        raise ValueError("Unsupported file format. Supported formats are CSV, XLS, XLSX, JSON, and TSV.")

# Function to load inbuilt dataset
def load_inbuilt_dataset(dataset_name):
    dataset_name = dataset_name.lower()
    if dataset_name == 'iris':
        return sns.load_dataset('iris')
    elif dataset_name == 'tips':
        return sns.load_dataset('tips')
    elif dataset_name == 'titanic':
        return sns.load_dataset('titanic')
    else:
        raise ValueError("Unsupported dataset.")

# Function for automated EDA
def automated_eda(data):
    # Summary statistics
    summary = data.describe()

    # Data types
    data_types = data.dtypes

    # Missing values
    missing_values = data.isnull().sum()

    # Exclude non-numeric columns from correlation matrix calculation
    numeric_columns = data.select_dtypes(include=['int64', 'float64']).columns
    correlation_matrix = data[numeric_columns].corr()

    # Distribution plots
    for column in numeric_columns:
        fig = go.Figure(data=[go.Histogram(x=data[column])])
        fig.update_layout(
            title=f'Distribution of {column}',
            xaxis_title=column,
            yaxis_title='Frequency'
        )
        st.plotly_chart(fig)

    # Correlation heatmap using Plotly
    if len(numeric_columns) >= 2:
        corr_data = data[numeric_columns].corr()
    
        # Create a heatmap using Plotly's figure_factory
        fig = ff.create_annotated_heatmap(
            z=corr_data.values,
            x=corr_data.columns,
            y=corr_data.index,
            colorscale='coolwarm',
            showscale=True
        )
    
        # Customize the layout
        fig.update_layout(
            title='Correlation Heatmap',
            xaxis_title='Columns',
            yaxis_title='Columns'
        )
    
        st.plotly_chart(fig)

    
    # Pairwise scatter plots (for numeric columns)
    if len(numeric_columns) >= 2:
        fig = px.scatter_matrix(data, dimensions=numeric_columns, title='Interactive Scatter Plot Matrix')
        st.plotly_chart(fig)
    
    # Box plots (for numeric columns) using Plotly
    for column in numeric_columns:
        fig = px.box(data, x=column, title=f'Box Plot of {column}')
        st.plotly_chart(fig)


    
    # Interactive scatter plot matrix (using Plotly)
    if len(numeric_columns) >= 2:
        fig = px.scatter_matrix(data, dimensions=numeric_columns, title='Interactive Scatter Plot Matrix')
        st.plotly_chart(fig)

    # Interactive histogram for numeric columns
    for column in numeric_columns:
        fig = px.histogram(data, x=column, title=f'Histogram of {column}')
        st.plotly_chart(fig)
        
    return summary, data_types, missing_values, correlation_matrix



# Define file_format variable
file_format = None

# Select dataset format
file_format = st.selectbox("Select a dataset format:", ["CSV", "XLS", "JSON", "TSV", "Inbuilt Datasets"])


if file_format == "Inbuilt Datasets":
    dataset_name = st.selectbox("Select an inbuilt dataset:", ["Iris", "Tips", "Titanic"])
    data = sns.load_dataset(dataset_name)
else:
    uploaded_file = st.file_uploader(f"Upload a {file_format} file", type=[file_format.lower()])
    if uploaded_file is not None:
        sheet_name = None
        if file_format in ["XLS", "XLSX"]:
            # Allow users to select a sheet from Excel files
            sheet_name = st.selectbox("Select a sheet (optional):", [""] + load_data(uploaded_file, file_format))
        data = load_data(uploaded_file, file_format, sheet_name=sheet_name)

if 'data' in locals():
    st.write("### Dataset Preview:")
    # Boolean to resize the dataframe, stored as a session state variable
    st.checkbox("Use container width", value=False, key="use_container_width")
    n, m = data.shape
    st.write(f'<p style="font-size:130%">Dataset contains {n} rows and {m} columns.</p>', unsafe_allow_html=True)
    st.dataframe(data, use_container_width=st.session_state.use_container_width)

    st.write("### Automated EDA:")
    summary, data_types, missing_values, correlation_matrix = automated_eda(data)

    st.write("#### Summary Statistics:")
    st.write(summary)

    st.write("#### Data Types:")
    st.write(data_types)

    st.write("#### Missing Values:")
    st.write(missing_values)

    st.write("#### Correlation Matrix:")
    st.write(correlation_matrix)

    
    data_info = ['Info', 'Null Info', 'Box Plots', 'Descriptive Analysis', 'Automated EDA']


    sdbar = st.sidebar.multiselect("EDA Options: ", data_info)

    if 'Info' in sdbar:
        st.subheader('Info:')
        c1, c2, c3 = st.columns([1, 2, 1])
        c2.dataframe(datf_inf(data))

    if 'Null Info' in sdbar:
        st.subheader('NA Value Information:')
        if data.isnull().sum().sum() == 0:
            st.write('There is not any NA value in your dataset.')
        else:
            c1, c2, c3 = st.columns([0.5, 2, 0.5])
            c2.dataframe(datf_nval(data), width=1500)
            st.markdown('')

    if 'Descriptive Analysis' in sdbar:
        st.subheader('Descriptive Analysis:')
        st.dataframe(data.describe())
    num_columns = data.select_dtypes(exclude='object').columns
    
    if 'Box Plots' in sdbar:
        if len(num_columns) == 0:
            st.write('There is no numerical columns in the data.')
        else:
            selected_num_cols = sidebar_multiselect_container('Choose columns for Box plots:', num_columns,
                                                                        'Box')
            st.subheader('Box plots')
            i = 0
            while (i < len(selected_num_cols)):
                c1, c2 = st.columns(2)
                for j in [c1, c2]:

                    if (i >= len(selected_num_cols)):
                        break

                    fig = px.box(data, y=selected_num_cols[i])
                    j.plotly_chart(fig, use_container_width=True)
                    i += 1

    if 'Automated EDA' in sdbar:
        datf = data
        st.write("Please Wait for Few Seconds.....")

        pr = data.profile_report(dark_mode=True)

        progress_bar = st.progress(0)
        status_text = st.empty()
        chart = st.line_chart(np.random.randn(10, 2))


        st_profile_report(pr)
        st.balloons()
st.markdown("<div id='linkto_top'></div>", unsafe_allow_html=True)


pages = ["Page 1","Page 2"]
section = st.sidebar.radio('', pages)

if section == "Page 1":                  # This is the beginning of my first page

    # add the link at the bottom of each page
    st.markdown("<a href='#linkto_top'>Link to top</a>", unsafe_allow_html=True)
